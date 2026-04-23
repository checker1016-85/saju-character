# -*- coding: utf-8 -*-
"""
사주명리_완전정리_DB.xlsx → saju_db.json 변환기
────────────────────────────────────────────────────────────
GitHub Actions 자동화용. DB 수정 후 재변환.

사용법:
  python excel_to_json.py                     # 기본 경로 사용
  python excel_to_json.py <xlsx_path>         # 특정 xlsx 경로 지정
  python excel_to_json.py <xlsx_path> <json_out>
                                              # 출력 경로까지 지정

동작:
  1. 엑셀의 각 시트를 파싱하여 JSON 구조로 변환
  2. 시트18 기반으로 시트19(일주별 직업군 매칭)를 자동 재생성
     → 시트18 편집만 하면 시트19는 매 실행마다 동기화됨
  3. saju_db.json 하나의 파일로 통합 출력

의존성:
  pip install pandas openpyxl

JSON 구조:
  {
    "dna": {...},                   # 시트12 — 60갑자 common/male/female DNA
    "gender_traits": {...},         # 시트10 — 천간×성별 특성
    "ilju_unique": {...},           # 시트11 — 60일주 고유 성격/외형
    "sipsung_jobs": {...},          # 시트6 — 십성별 직업방향
    "job_visual": {...},            # 시트14 — 직업별 비주얼 DB
    "body_pct": {...},              # 시트15 — 60일주 × 7카테고리 신체 퍼센트
    "ilju_combo": {...},            # 시트16 — 조합 키워드/격국/설명
    "month_correction": {...},      # 시트17 — 12월지 보정
    "branch_harmony": {...},        # 시트17 하단 — 지지 합충 관계표
    "job_categories_100": [...],    # 시트18 — 100 직업군(분리 태그 포함)
    "job_exclude_option": {...},    # 시트18 JC000 행
    "ilju_to_jobs": {...},          # 시트19 자동 재생성 — 일주 역인덱스
    "meta": {...}                   # 빌드 메타데이터
  }
"""
import os
import sys
import json
from collections import defaultdict
from datetime import datetime

import pandas as pd
import openpyxl


# ─────────────────────────────────────────────────────────────
# 경로 처리
# ─────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(SCRIPT_DIR, '사주명리_완전정리_DB.xlsx')
DEFAULT_JSON = os.path.join(SCRIPT_DIR, 'saju_db.json')


# ─────────────────────────────────────────────────────────────
# 공통 유틸
# ─────────────────────────────────────────────────────────────
def cell(df, r, c):
    """pandas DataFrame의 (r, c) 셀을 안전하게 문자열로"""
    if r >= len(df) or c >= len(df.columns):
        return ''
    v = df.iloc[r, c]
    return str(v).strip() if pd.notna(v) and str(v).strip() else ''


def split_list(s, seps=(',', '·', '/')):
    """문자열을 여러 구분자로 잘라 공백 제거 후 반환"""
    if not s:
        return []
    tokens = [s]
    for sep in seps:
        new = []
        for t in tokens:
            new.extend(t.split(sep))
        tokens = new
    return [t.strip() for t in tokens if t.strip() and t.strip() != '—']


# ─────────────────────────────────────────────────────────────
# 시트12 — 일주 캐릭터 DNA
# ─────────────────────────────────────────────────────────────
COL12 = {
    4: 'A1_골격', 5: 'A2_체형',
    6: 'B1_얼굴형', 7: 'B2_눈', 8: 'B3_이면눈매',
    9: 'B4_눈썹', 10: 'B5_코', 11: 'B6_입', 12: 'B7_얼굴음영',
    13: 'C1_피부톤', 14: 'C2_오행변이',
    15: 'D1_헤어', 16: 'D2_머리물리', 17: 'D3_수염',
    18: 'E1_의복핏', 19: 'E2_상의', 20: 'E3_하의', 21: 'E4_풀바디', 22: 'E5_신발',
    23: 'F1_머리악세', 24: 'F2_몸악세', 25: 'F3_소울웨폰', 26: 'F4_장신구공명',
    27: 'F5_인장문양', 28: 'F6_특수요소',
    29: 'G1_아우라', 30: 'G2_이펙트컬러', 31: 'G3_그림자', 32: 'G4_발걸음',
    33: 'G5_동작잔상', 34: 'G6_지장간',
    35: 'H1_표정온도', 36: 'H2_분위기', 37: 'H3_제스처', 38: 'H4_음성톤',
    39: 'H5_감정중력', 40: 'H6_각성트리거', 41: 'H7_나이대종족',
}


def extract_dna(xlsx):
    df = pd.read_excel(xlsx, sheet_name='시트12_일주_캐릭터DNA', header=None)
    dna = {}
    cur = None
    for i in range(3, len(df)):
        ilju = cell(df, i, 1)
        if ilju:
            cur = ilju
        if not cur:
            continue
        g = {'共': 'common', '♂': 'male', '♀': 'female'}.get(cell(df, i, 3))
        if not g:
            continue
        if cur not in dna:
            dna[cur] = {}
        rd = {}
        for ci, cn in COL12.items():
            v = cell(df, i, ci)
            if v:
                rd[cn] = v
        dna[cur][g] = rd
    return dna


# ─────────────────────────────────────────────────────────────
# 시트10 — 천간×성별 특성
# ─────────────────────────────────────────────────────────────
def extract_gender(xlsx):
    df = pd.read_excel(xlsx, sheet_name='시트10_성별', header=None)
    gt = {}
    cols = {
        3: '성격분기', 4: '외형분기', 5: '체형분기', 6: '피부인상분기',
        7: '눈빛표정분기', 8: '분위기태그', 9: '스타일분기',
        10: '직업방향', 11: '관계패턴', 12: '프롬프트외형태그',
    }
    for i in range(2, len(df)):
        cg = cell(df, i, 1)
        gd = cell(df, i, 2)
        if not cg or not gd:
            continue
        hj = cg.split('(')[1].split(')')[0] if '(' in cg else cg
        key = f"{hj}_{gd}"
        rd = {}
        for ci, cn in cols.items():
            v = cell(df, i, ci)
            if v:
                rd[cn] = v
        gt[key] = rd
    return gt


# ─────────────────────────────────────────────────────────────
# 시트11 — 60일주 고유 특성
# ─────────────────────────────────────────────────────────────
def extract_ilju(xlsx):
    df = pd.read_excel(xlsx, sheet_name='시트11_일주', header=None)
    iu = {}
    cols = {
        3: '천간', 4: '지지', 5: '납음', 6: '납음해석', 7: '십이운성',
        8: '일지십성', 9: '생극관계', 10: '고유성격', 11: '고유외형', 12: '고유이미지',
    }
    for i in range(2, len(df)):
        ilju = cell(df, i, 2)
        if not ilju:
            continue
        rd = {}
        for ci, cn in cols.items():
            v = cell(df, i, ci)
            if v:
                rd[cn] = v
        iu[ilju] = rd
    return iu


# ─────────────────────────────────────────────────────────────
# 시트6 — 십성별 직업방향
# ─────────────────────────────────────────────────────────────
def extract_sipsung_jobs(xlsx):
    df = pd.read_excel(xlsx, sheet_name='시트6_지지십성', header=None)
    sj = {}
    for i in range(2, len(df)):
        name = cell(df, i, 1)
        job = cell(df, i, 6)
        if name and job:
            sj[name] = job
    return sj


# ─────────────────────────────────────────────────────────────
# 시트14 — 직업별 비주얼
# ─────────────────────────────────────────────────────────────
def extract_job_visual(xlsx):
    try:
        df = pd.read_excel(xlsx, sheet_name='시트14_직업비주얼', header=None)
    except Exception:
        return {}
    jv = {}
    for i in range(2, len(df)):
        name = cell(df, i, 1)
        desc = cell(df, i, 2)
        sd = cell(df, i, 3)
        if name:
            entry = {}
            if desc:
                entry['desc'] = desc
            if sd:
                entry['sd'] = sd
            jv[name] = entry
    return jv


# ─────────────────────────────────────────────────────────────
# 시트15 — 60일주 신체 퍼센트 매트릭스
# ─────────────────────────────────────────────────────────────
# 23개 퍼센트 컬럼 (C~Y)을 7개 카테고리로 묶음
BODY_LAYOUT = [
    ('키',       4,  ['단신', '중간', '장신', '매우장신']),
    ('체형',     4,  ['마른', '보통', '근육', '풍채']),
    ('나잇대',   3,  ['동안', '또래', '성숙']),
    ('얼굴',     3,  ['작은편', '보통', '큰편']),
    ('이목구비', 3,  ['섬세', '보통', '뚜렷']),
    ('피부',     3,  ['백옥', '건강', '구릿빛']),
    ('모발',     3,  ['가늘', '보통', '굵은']),
]


def extract_body_pct(xlsx):
    """
    시트15 구조:
      - 헤더 2개 행 (병합 대분류 + 하위 세부)
      - 4번째 행부터 데이터 (A: 번호, B: 일주, C~: 퍼센트)
    """
    try:
        df = pd.read_excel(xlsx, sheet_name='시트15_일주신체분포', header=None)
    except Exception:
        return {}

    body = {}
    for i in range(3, len(df)):
        ilju = cell(df, i, 1)
        if not ilju:
            continue
        row_data = {}
        col_idx = 2  # C열부터 시작
        for cat_name, n, sublabels in BODY_LAYOUT:
            pcts = []
            for k in range(n):
                v = df.iloc[i, col_idx + k] if col_idx + k < len(df.columns) else 0
                try:
                    pcts.append(int(v) if pd.notna(v) else 0)
                except (ValueError, TypeError):
                    pcts.append(0)
            row_data[cat_name] = pcts
            col_idx += n
        body[ilju] = row_data
    return body


# ─────────────────────────────────────────────────────────────
# 시트16 — 60일주 조합 키워드
# ─────────────────────────────────────────────────────────────
def extract_ilju_combo(xlsx):
    try:
        df = pd.read_excel(xlsx, sheet_name='시트16_일주조합키워드', header=None)
    except Exception:
        return {}

    combo = {}
    for i in range(2, len(df)):
        ilju = cell(df, i, 1)
        if not ilju:
            continue
        combo[ilju] = {
            'kw_ko':      split_list(cell(df, i, 2), seps=(' · ',)),
            'kw_en':      [x.strip() for x in cell(df, i, 3).split(',') if x.strip()],
            '격국':       cell(df, i, 4),
            '특수격':     cell(df, i, 5),
            '오행강약':   cell(df, i, 6),
            '설명':       cell(df, i, 7),
        }
    return combo


# ─────────────────────────────────────────────────────────────
# 시트17 — 월지 보정 + 지지 합충 관계표
# ─────────────────────────────────────────────────────────────
def extract_month_correction(xlsx):
    """월지 보정표(상단 12행) + 합충 관계표(하단)를 분리 반환"""
    try:
        df = pd.read_excel(xlsx, sheet_name='시트17_월지보정', header=None)
    except Exception:
        return {}, {}

    month = {}
    # 상단 12행: 월지 보정
    for i in range(2, len(df)):
        br = cell(df, i, 0)
        # 지지 1글자 + 유효 지지만 허용
        if len(br) != 1 or br not in '寅卯辰巳午未申酉戌亥子丑':
            continue
        if len(month) >= 12:
            break
        month[br] = {
            '계절':       cell(df, i, 1),
            '오행':       cell(df, i, 2),
            '조후':       cell(df, i, 3),
            '장간':       cell(df, i, 4),
            '월령십성':   cell(df, i, 5),
            '키워드':     split_list(cell(df, i, 6), seps=(' · ',)),
            '설명':       cell(df, i, 7),
        }

    # 하단 합충 관계표: 2글자 지지 조합
    harmony = {}
    for i in range(2, len(df)):
        a = cell(df, i, 0)
        b = cell(df, i, 1)
        rel = cell(df, i, 2)
        if (len(a) == 1 and len(b) == 1 and a != b
                and a in '寅卯辰巳午未申酉戌亥子丑'
                and b in '寅卯辰巳午未申酉戌亥子丑'
                and rel):
            harmony[f'{a}_{b}'] = rel

    return month, harmony


# ─────────────────────────────────────────────────────────────
# 시트18 — 직업군 100 (분리 태그 포함)
# ─────────────────────────────────────────────────────────────
def extract_jobs_100(xlsx):
    """
    시트18 구조 (10개 컬럼):
      A 코드 / B 직업군명 / C 대분류 / D 키워드 / E 오행 태그 /
      F 십성 태그 / G 기질 태그 / H 특수격 태그 / I 매칭 일주 수 /
      J 매칭 일주 목록
    """
    try:
        df = pd.read_excel(xlsx, sheet_name='시트18_직업군100', header=None)
    except Exception:
        return [], None

    jobs = []
    jc000 = None
    for i in range(2, len(df)):
        code = cell(df, i, 0)
        if not code or not code.startswith('JC'):
            continue

        matching = [
            p.strip() for p in cell(df, i, 9).split(',')
            if p.strip() and p.strip() not in ('—', '전체 60일주 선택 가능')
        ]
        keywords = split_list(cell(df, i, 3), seps=(' · ',))

        entry = {
            'id':            code,
            'name':          cell(df, i, 1),
            'category':      cell(df, i, 2),
            'keywords':      keywords,
            'oheng':         cell(df, i, 4),
            'sipsung':       cell(df, i, 5),
            'gijil':         cell(df, i, 6),
            'teukgyeok':     cell(df, i, 7),
            'matching_ilju': matching,
        }

        if code == 'JC000':
            jc000 = entry
        else:
            jobs.append(entry)
    return jobs, jc000


# ─────────────────────────────────────────────────────────────
# 시트19 자동 재생성
#   시트18의 매칭 일주 목록을 스캔하여 역인덱스 계산
#   → 엑셀 시트19에 쓰기 + 반환값은 JSON용
# ─────────────────────────────────────────────────────────────
def refresh_sheet19(xlsx, jobs):
    """시트19 자동 갱신. 반환: {일주: [JC코드, ...]}"""
    ilju_to_jobs = defaultdict(list)
    for j in jobs:
        for p in j['matching_ilju']:
            ilju_to_jobs[p].append(j['id'])
    ilju_to_jobs = dict(ilju_to_jobs)

    # 엑셀 시트19에 값 기입 (openpyxl)
    try:
        wb = openpyxl.load_workbook(xlsx)
        if '시트19_일주별직업군매칭' in wb.sheetnames:
            ws = wb['시트19_일주별직업군매칭']
            # 헤더 3행 아래 데이터는 r=4부터 (60일주)
            for r in range(4, 64):
                ilju = ws.cell(row=r, column=2).value
                if not ilju:
                    continue
                codes = ilju_to_jobs.get(ilju, [])
                ws.cell(row=r, column=3).value = len(codes)
                ws.cell(row=r, column=4).value = ', '.join(codes) if codes else ''
            wb.save(xlsx)
    except Exception as e:
        # 엑셀 쓰기 실패해도 JSON은 정상 반환 (CI 환경 등에서 파일 잠김 대비)
        print(f'  WARN: 시트19 엑셀 쓰기 실패 ({e}). JSON은 정상 생성됨.')

    return ilju_to_jobs


# ─────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────
def main(xlsx=DEFAULT_XLSX, json_out=DEFAULT_JSON):
    print(f'Reading {xlsx}...')

    if not os.path.exists(xlsx):
        print(f'ERROR: {xlsx} 파일이 없습니다.')
        sys.exit(1)

    # 모든 시트 추출
    month, harmony = extract_month_correction(xlsx)
    jobs_100, jc000 = extract_jobs_100(xlsx)
    ilju_to_jobs = refresh_sheet19(xlsx, jobs_100)

    result = {
        'dna':                 extract_dna(xlsx),
        'gender_traits':       extract_gender(xlsx),
        'ilju_unique':         extract_ilju(xlsx),
        'sipsung_jobs':        extract_sipsung_jobs(xlsx),
        'job_visual':          extract_job_visual(xlsx),
        'body_pct':            extract_body_pct(xlsx),
        'ilju_combo':          extract_ilju_combo(xlsx),
        'month_correction':    month,
        'branch_harmony':      harmony,
        'job_categories_100':  jobs_100,
        'job_exclude_option':  jc000,
        'ilju_to_jobs':        ilju_to_jobs,
        'meta': {
            'schema_version': '2.0',
            'generated_at':   datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'source':         os.path.basename(xlsx),
        },
    }

    with open(json_out, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, separators=(',', ':'))

    sz = os.path.getsize(json_out) / 1024
    print(f'Done -> {json_out} ({sz:.0f}KB)')
    print(f'  DNA:           {len(result["dna"])}')
    print(f'  Gender:        {len(result["gender_traits"])}')
    print(f'  Ilju:          {len(result["ilju_unique"])}')
    print(f'  SipsungJobs:   {len(result["sipsung_jobs"])}')
    print(f'  JobVisual:     {len(result["job_visual"])}')
    print(f'  BodyPct:       {len(result["body_pct"])}')
    print(f'  IljuCombo:     {len(result["ilju_combo"])}')
    print(f'  MonthCorr:     {len(result["month_correction"])}')
    print(f'  Harmony:       {len(result["branch_harmony"])}')
    print(f'  Jobs100:       {len(result["job_categories_100"])}')
    print(f'  IljuToJobs:    {len(result["ilju_to_jobs"])} iljus indexed')


if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    out  = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_JSON
    main(xlsx, out)
